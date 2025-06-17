"""
Excel Generation Module
Creates formatted Excel reports from processed inquiry data
"""

import os
from datetime import datetime
from typing import List, Dict, Any
import pandas as pd
from openpyxl import Workbook, load_workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils.dataframe import dataframe_to_rows

from config import Config
from utils.logger import setup_logger

class ExcelGenerator:
    """Generates Excel reports from processed customer inquiry data"""
    
    def __init__(self):
        self.config = Config()
        self.logger = setup_logger('excel_generator')
        
        # Excel styling
        self.header_font = Font(bold=True, color='FFFFFF')
        self.header_fill = PatternFill(start_color='366092', end_color='366092', fill_type='solid')
        self.border = Border(
            left=Side(style='thin'),
            right=Side(style='thin'),
            top=Side(style='thin'),
            bottom=Side(style='thin')
        )
        self.center_alignment = Alignment(horizontal='center', vertical='center')
    
    def create_dataframe(self, results: List[Dict[str, Any]]) -> pd.DataFrame:
        """
        Create pandas DataFrame from processing results
        
        Args:
            results: List of processing results
            
        Returns:
            Formatted DataFrame
        """
        try:
            # Prepare data for DataFrame
            data_rows = []
            
            for result in results:
                # Extract file name from path
                file_name = os.path.basename(result.get('file_path', ''))
                
                # Determine extraction method
                extraction_method = self._get_primary_extraction_method(result)
                
                row = {
                    'File Name': file_name,
                    'Customer Name': result.get('customer_name', ''),
                    'Travel Dates': result.get('travel_dates', ''),
                    'Destination': result.get('destination', ''),
                    'Budget': result.get('budget', ''),
                    'Number of Travelers': result.get('travelers_count', ''),
                    'Contact Information': result.get('contact_info', ''),
                    'Special Requirements': result.get('special_requirements', ''),
                    'Processing Status': result.get('status', 'PROCESSED'),
                    'Processing Time (seconds)': round(result.get('processing_time', 0.0), 3),
                    'Extraction Method': extraction_method,
                    'Confidence Score': round(result.get('confidence_score', 0.0), 2) if result.get('confidence_score') else 'N/A'
                }
                
                data_rows.append(row)
            
            # Create DataFrame
            df = pd.DataFrame(data_rows)
            
            # Ensure all expected columns are present
            for col in self.config.EXCEL_COLUMNS:
                if col not in df.columns:
                    df[col] = ''
            
            # Reorder columns according to config
            df = df[self.config.EXCEL_COLUMNS + ['Confidence Score']]
            
            self.logger.info(f"Created DataFrame with {len(df)} rows and {len(df.columns)} columns")
            return df
            
        except Exception as e:
            self.logger.error(f"DataFrame creation error: {str(e)}")
            # Return empty DataFrame with correct columns
            return pd.DataFrame(columns=self.config.EXCEL_COLUMNS + ['Confidence Score'])
    
    def _get_primary_extraction_method(self, result: Dict[str, Any]) -> str:
        """
        Determine primary extraction method used
        
        Args:
            result: Processing result dictionary
            
        Returns:
            Primary extraction method string
        """
        if 'extraction_methods' in result:
            methods = result['extraction_methods']
            
            # Count method occurrences
            method_counts = {}
            for method in methods.values():
                method_counts[method] = method_counts.get(method, 0) + 1
            
            # Return most common method
            if method_counts:
                primary_method = max(method_counts, key=method_counts.get)
                if method_counts.get('COMBINED', 0) > 0:
                    return 'HYBRID (ML + Rules)'
                elif primary_method == 'ML_NER':
                    return 'ML (BERT NER)'
                elif primary_method == 'RULE_BASED':
                    return 'Rule-Based'
                else:
                    return 'Unknown'
        
        return 'Standard Processing'
    
    def style_worksheet(self, worksheet, df: pd.DataFrame):
        """
        Apply styling to the worksheet
        
        Args:
            worksheet: openpyxl worksheet object
            df: DataFrame containing the data
        """
        try:
            # Style header row
            for col_num, cell in enumerate(worksheet[1], 1):
                cell.font = self.header_font
                cell.fill = self.header_fill
                cell.alignment = self.center_alignment
                cell.border = self.border
            
            # Style data rows
            for row in worksheet.iter_rows(min_row=2, max_row=len(df) + 1):
                for cell in row:
                    cell.border = self.border
                    cell.alignment = Alignment(horizontal='left', vertical='center', wrap_text=True)
            
            # Adjust column widths
            column_widths = {
                'A': 20,  # File Name
                'B': 25,  # Customer Name
                'C': 30,  # Travel Dates
                'D': 20,  # Destination
                'E': 15,  # Budget
                'F': 15,  # Number of Travelers
                'G': 35,  # Contact Information
                'H': 30,  # Special Requirements
                'I': 20,  # Processing Status
                'J': 18,  # Processing Time
                'K': 20,  # Extraction Method
                'L': 15,  # Confidence Score
            }
            
            for col_letter, width in column_widths.items():
                worksheet.column_dimensions[col_letter].width = width
            
            # Set row height for header
            worksheet.row_dimensions[1].height = 25
            
            self.logger.debug("Worksheet styling applied successfully")
            
        except Exception as e:
            self.logger.error(f"Worksheet styling error: {str(e)}")
    
    def add_summary_sheet(self, workbook: Workbook, results: List[Dict[str, Any]]):
        """
        Add a summary sheet with statistics
        
        Args:
            workbook: openpyxl workbook object
            results: List of processing results
        """
        try:
            # Create summary worksheet
            summary_ws = workbook.create_sheet("Summary", 0)
            
            # Calculate statistics
            total_files = len(results)
            successful = sum(1 for r in results if not r.get('status', '').startswith('ERROR'))
            errors = total_files - successful
            
            # Processing time statistics
            processing_times = [r.get('processing_time', 0.0) for r in results]
            avg_time = sum(processing_times) / len(processing_times) if processing_times else 0
            total_time = sum(processing_times)
            
            # Extraction method statistics
            method_counts = {}
            for result in results:
                method = self._get_primary_extraction_method(result)
                method_counts[method] = method_counts.get(method, 0) + 1
            
            # Add summary data
            summary_data = [
                ['AI Travel Agent - Processing Summary', ''],
                ['Generated on:', datetime.now().strftime('%Y-%m-%d %H:%M:%S')],
                ['', ''],
                ['Processing Statistics', ''],
                ['Total Files Processed:', total_files],
                ['Successful Extractions:', successful],
                ['Failed Extractions:', errors],
                ['Success Rate:', f"{(successful/total_files)*100:.1f}%" if total_files > 0 else "0%"],
                ['', ''],
                ['Performance Statistics', ''],
                ['Total Processing Time:', f"{total_time:.2f} seconds"],
                ['Average Time per File:', f"{avg_time:.3f} seconds"],
                ['Files per Second:', f"{total_files/total_time:.2f}" if total_time > 0 else "N/A"],
                ['', ''],
                ['Extraction Methods Used', ''],
            ]
            
            # Add method statistics
            for method, count in method_counts.items():
                summary_data.append([method, f"{count} files ({(count/total_files)*100:.1f}%)"])
            
            # Write summary data
            for row_idx, (label, value) in enumerate(summary_data, 1):
                summary_ws.cell(row=row_idx, column=1, value=label)
                summary_ws.cell(row=row_idx, column=2, value=value)
            
            # Style summary sheet
            summary_ws.cell(row=1, column=1).font = Font(bold=True, size=16)
            summary_ws.cell(row=4, column=1).font = Font(bold=True, size=12)
            summary_ws.cell(row=10, column=1).font = Font(bold=True, size=12)
            summary_ws.cell(row=15, column=1).font = Font(bold=True, size=12)
            
            # Adjust column widths
            summary_ws.column_dimensions['A'].width = 30
            summary_ws.column_dimensions['B'].width = 20
            
            self.logger.info("Summary sheet added successfully")
            
        except Exception as e:
            self.logger.error(f"Summary sheet creation error: {str(e)}")
    
    def generate_report(self, results: List[Dict[str, Any]], output_path: str) -> str:
        """
        Generate Excel report from processing results
        
        Args:
            results: List of processing results
            output_path: Path for output Excel file
            
        Returns:
            Path to generated Excel file
        """
        try:
            self.logger.info(f"Generating Excel report: {output_path}")
            
            # Create DataFrame
            df = self.create_dataframe(results)
            
            if df.empty:
                self.logger.warning("No data to write to Excel")
                return self._create_empty_report(output_path)
            
            # Create Excel workbook
            workbook = Workbook()
            
            # Remove default sheet
            default_sheet = workbook.active
            workbook.remove(default_sheet)
            
            # Add summary sheet
            self.add_summary_sheet(workbook, results)
            
            # Create main data sheet
            data_ws = workbook.create_sheet("Inquiry Data")
            
            # Write DataFrame to worksheet
            for row in dataframe_to_rows(df, index=False, header=True):
                data_ws.append(row)
            
            # Style the worksheet
            self.style_worksheet(data_ws, df)
            
            # Ensure output directory exists
            os.makedirs(os.path.dirname(output_path), exist_ok=True)
            
            # Save workbook
            workbook.save(output_path)
            
            self.logger.info(f"Excel report generated successfully: {output_path}")
            return output_path
            
        except Exception as e:
            self.logger.error(f"Excel generation error: {str(e)}")
            return self._create_empty_report(output_path)
    
    def _create_empty_report(self, output_path: str) -> str:
        """
        Create an empty report when no data is available
        
        Args:
            output_path: Path for output Excel file
            
        Returns:
            Path to generated Excel file
        """
        try:
            # Create empty DataFrame with correct columns
            df = pd.DataFrame(columns=self.config.EXCEL_COLUMNS + ['Confidence Score'])
            
            # Add a message row
            df.loc[0] = ['No data available'] + [''] * (len(df.columns) - 1)
            
            # Create workbook and save
            workbook = Workbook()
            ws = workbook.active
            ws.title = "Inquiry Data"
            
            # Write empty DataFrame
            for row in dataframe_to_rows(df, index=False, header=True):
                ws.append(row)
            
            # Basic styling
            for cell in ws[1]:
                cell.font = self.header_font
                cell.fill = self.header_fill
            
            # Ensure output directory exists
            os.makedirs(os.path.dirname(output_path), exist_ok=True)
            
            workbook.save(output_path)
            
            self.logger.info(f"Empty Excel report created: {output_path}")
            return output_path
            
        except Exception as e:
            self.logger.error(f"Empty report creation error: {str(e)}")
            raise
    
    def validate_excel_file(self, file_path: str) -> bool:
        """
        Validate that the generated Excel file is readable
        
        Args:
            file_path: Path to Excel file
            
        Returns:
            True if file is valid, False otherwise
        """
        try:
            # Try to load the file
            workbook = load_workbook(file_path)
            
            # Check if it has the expected sheets
            expected_sheets = ['Summary', 'Inquiry Data']
            for sheet_name in expected_sheets:
                if sheet_name not in workbook.sheetnames:
                    self.logger.warning(f"Expected sheet '{sheet_name}' not found")
                    return False
            
            # Check if data sheet has content
            data_sheet = workbook['Inquiry Data']
            if data_sheet.max_row < 2:  # Should have at least header + 1 row
                self.logger.warning("Data sheet appears to be empty")
            
            workbook.close()
            return True
            
        except Exception as e:
            self.logger.error(f"Excel validation error: {str(e)}")
            return False
    
    def generate_schema_report(self, results: List[Dict[str, Any]], output_path: str) -> str:
        """
        Generate Excel report from schema-validated processing results
        
        Args:
            results: List of schema-validated processing results
            output_path: Path for output Excel file
            
        Returns:
            Path to generated Excel file
        """
        if not results:
            return self._create_empty_report(output_path)
        
        try:
            self.logger.info(f"Generating schema-based Excel report: {output_path}")
            
            # Create DataFrame with schema fields
            df = self.create_schema_dataframe(results)
            self.logger.info(f"Created DataFrame with {len(df)} rows and {len(df.columns)} columns")
            
            # Create Excel workbook
            with pd.ExcelWriter(output_path, engine='openpyxl') as writer:
                # Write main data sheet
                df.to_excel(writer, sheet_name='Travel Inquiries', index=False)
                
                # Style the worksheet
                workbook = writer.book
                worksheet = writer.sheets['Travel Inquiries']
                self.style_schema_worksheet(worksheet, df)
                
                # Add summary sheet
                self.add_schema_summary_sheet(workbook, results)
            
            self.logger.info(f"Excel report generated successfully: {output_path}")
            return output_path
                
        except Exception as e:
            self.logger.error(f"Error generating schema Excel report: {str(e)}")
            raise
    
    def create_schema_dataframe(self, results: List[Dict[str, Any]]) -> pd.DataFrame:
        """
        Create pandas DataFrame from schema-validated processing results
        
        Args:
            results: List of schema-validated processing results
            
        Returns:
            Formatted DataFrame with schema fields
        """
        # Define schema-based columns
        schema_columns = [
            'File Name', 'Customer Name', 'Contact Info',
            'Number of Travelers', 'Adults', 'Children',
            'Destinations', 'Start Date', 'End Date', 'Duration (Nights)',
            'Hotel Preferences', 'Meal Preferences', 'Activities',
            'Needs Flight', 'Budget', 'Special Requests',
            'Processing Status', 'Processing Time (seconds)', 'Extraction Method'
        ]
        
        # Convert results to DataFrame rows
        data_rows = []
        for result in results:
            row = {
                'File Name': result.get('file_name', ''),
                'Customer Name': result.get('customer_name', ''),
                'Contact Info': result.get('contact_info', ''),
                'Number of Travelers': result.get('num_travelers', ''),
                'Adults': result.get('num_adults', ''),
                'Children': result.get('num_children', ''),
                'Destinations': ', '.join(result.get('destinations', [])) if result.get('destinations') else '',
                'Start Date': result.get('start_date', ''),
                'End Date': result.get('end_date', ''),
                'Duration (Nights)': result.get('duration_nights', ''),
                'Hotel Preferences': result.get('hotel_preferences', ''),
                'Meal Preferences': result.get('meal_preferences', ''),
                'Activities': ', '.join(result.get('activities', [])) if result.get('activities') else '',
                'Needs Flight': 'Yes' if result.get('needs_flight') is True else 'No' if result.get('needs_flight') is False else '',
                'Budget': result.get('budget', ''),
                'Special Requests': result.get('special_requests', ''),
                'Processing Status': result.get('processing_status', ''),
                'Processing Time (seconds)': f"{result.get('processing_time', 0):.3f}",
                'Extraction Method': result.get('extraction_method', '')
            }
            data_rows.append(row)
        
        # Create DataFrame
        df = pd.DataFrame(data_rows, columns=schema_columns)
        
        # Sort by file name for consistent ordering
        df = df.sort_values('File Name').reset_index(drop=True)
        
        return df
    
    def style_schema_worksheet(self, worksheet, df: pd.DataFrame):
        """
        Apply styling to the schema-based worksheet
        
        Args:
            worksheet: openpyxl worksheet object
            df: DataFrame containing the data
        """
        try:
            from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
            
            # Header styling
            header_font = Font(bold=True, color="FFFFFF")
            header_fill = PatternFill(start_color="366092", end_color="366092", fill_type="solid")
            header_alignment = Alignment(horizontal="center", vertical="center")
            
            # Apply header styling
            for col_num, column in enumerate(df.columns, 1):
                cell = worksheet.cell(row=1, column=col_num)
                cell.font = header_font
                cell.fill = header_fill
                cell.alignment = header_alignment
            
            # Auto-adjust column widths
            for column in worksheet.columns:
                max_length = 0
                column_letter = column[0].column_letter
                
                for cell in column:
                    try:
                        if len(str(cell.value)) > max_length:
                            max_length = len(str(cell.value))
                    except:
                        pass
                
                adjusted_width = min(max_length + 2, 50)  # Cap at 50 characters
                worksheet.column_dimensions[column_letter].width = adjusted_width
            
            # Freeze header row
            worksheet.freeze_panes = "A2"
            
        except Exception as e:
            self.logger.warning(f"Error applying worksheet styling: {str(e)}")
    
    def add_schema_summary_sheet(self, workbook, results: List[Dict[str, Any]]):
        """
        Add a summary sheet with schema-based statistics
        
        Args:
            workbook: openpyxl workbook object
            results: List of schema-validated processing results
        """
        try:
            from openpyxl.styles import Font, PatternFill
            
            summary_sheet = workbook.create_sheet(title="Summary")
            
            # Calculate statistics
            total_files = len(results)
            successful = len([r for r in results if r.get('processing_status') == 'completed'])
            errors = total_files - successful
            
            # Language distribution
            language_stats = {}
            for result in results:
                file_name = result.get('file_name', '')
                if file_name.startswith('hindi_'):
                    lang = 'Hindi'
                elif file_name.startswith('english_'):
                    lang = 'English'
                elif file_name.startswith('hinglish_'):
                    lang = 'Hinglish'
                elif file_name.startswith('hindi_english_'):
                    lang = 'Hindi-English'
                else:
                    lang = 'Other'
                language_stats[lang] = language_stats.get(lang, 0) + 1
            
            # Write summary data
            header_font = Font(bold=True, size=14, color="FFFFFF")
            header_fill = PatternFill(start_color="366092", end_color="366092", fill_type="solid")
            
            # Main statistics
            summary_sheet['A1'] = "Travel Inquiries Processing Summary"
            summary_sheet['A1'].font = header_font
            summary_sheet['A1'].fill = header_fill
            
            row = 3
            summary_sheet[f'A{row}'] = "Total Files Processed:"
            summary_sheet[f'B{row}'] = total_files
            row += 1
            
            summary_sheet[f'A{row}'] = "Successful Extractions:"
            summary_sheet[f'B{row}'] = successful
            row += 1
            
            summary_sheet[f'A{row}'] = "Errors:"
            summary_sheet[f'B{row}'] = errors
            row += 2
            
            # Language distribution
            summary_sheet[f'A{row}'] = "Language Distribution:"
            summary_sheet[f'A{row}'].font = Font(bold=True)
            row += 1
            
            for lang, count in language_stats.items():
                summary_sheet[f'A{row}'] = f"  {lang}:"
                summary_sheet[f'B{row}'] = count
                row += 1
            
            self.logger.info("Schema summary sheet added successfully")
            
        except Exception as e:
            self.logger.error(f"Error creating schema summary sheet: {str(e)}")
